#!/usr/bin/env python3
"""Build a working spreadsheet from the "דרך קיצור" Google My Maps layer.

The map is the live source of truth for the initiative and changes often, so
this is a regenerator, not a one-off export: re-run it and the sheet is current.

Usage:
    python3 kitzur_sheet.py [output.xlsx]

Sheets produced:
    מקטעים      one row per trail segment, sorted longest first
    נקודות ציון  one row per waypoint
    סיכום        totals and the length distribution
"""
import html
import math
import re
import sys
import xml.etree.ElementTree as ET

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

MID = "19q4SgSScTh3pW3lRqQO9dQ8kw8z2wTg"
KML_URL = f"https://www.google.com/maps/d/kml?mid={MID}&forcekml=1"
MAP_URL = f"https://www.google.com/maps/d/viewer?hl=iw&mid={MID}"
NS = {"k": "http://www.opengis.net/kml/2.2"}
H = {"User-Agent": "Pardespedia-Bot/1.0 (orimosenzon@gmail.com)"}


def haversine(a, b):
    """Metres between two (lon, lat) pairs."""
    r = 6371000.0
    la1, la2 = math.radians(a[1]), math.radians(b[1])
    dla = la2 - la1
    dlo = math.radians(b[0] - a[0])
    h = math.sin(dla / 2) ** 2 + math.cos(la1) * math.cos(la2) * math.sin(dlo / 2) ** 2
    return 2 * r * math.asin(math.sqrt(h))


def coords(el):
    txt = el.findtext("k:coordinates", default="", namespaces=NS).strip()
    out = []
    for tok in txt.split():
        parts = tok.split(",")
        if len(parts) >= 2:
            out.append((float(parts[0]), float(parts[1])))
    return out


def clean(desc_raw):
    """Strip the HTML My Maps wraps photo captions in, keep the human note."""
    text = re.sub(r"<[^>]+>", " ", desc_raw)
    return re.sub(r"\s+", " ", html.unescape(text)).strip()


def parse(kml_bytes):
    root = ET.fromstring(kml_bytes)
    rows = []
    for pm in root.findall(".//k:Placemark", NS):
        name = (pm.findtext("k:name", default="", namespaces=NS) or "").strip()
        desc_raw = pm.findtext("k:description", default="", namespaces=NS) or ""
        line = pm.find(".//k:LineString", NS)
        point = pm.find(".//k:Point", NS)
        if line is not None:
            pts = coords(line)
            length = sum(haversine(pts[i], pts[i + 1]) for i in range(len(pts) - 1))
            kind = "מקטע שביל"
        elif point is not None:
            pts = coords(point)
            length = 0.0
            kind = "נקודת ציון"
        else:
            continue
        if not pts:
            continue
        start, end = pts[0], pts[-1]
        beeline = haversine(start, end)
        rows.append(
            {
                "שם": name,
                "סוג": kind,
                "אורך (מ')": round(length),
                "מרחק אווירי (מ')": round(beeline),
                "פיתולים": round(length / beeline, 2) if beeline > 1 else "",
                "נק' בקו": len(pts),
                "הערה מהמפה": clean(desc_raw),
                "תמונות": len(re.findall(r"<img\s", desc_raw)),
                "רחובות בשם": " · ".join(streets(name)),
                "קו רוחב": round(start[1], 6),
                "קו אורך": round(start[0], 6),
                "ניווט": f"https://www.google.com/maps/search/?api=1&query={start[1]:.6f},{start[0]:.6f}&hl=iw",
            }
        )
    return rows


SPLIT = re.compile(r"\s*[-–—]\s*|\s+ו\s+")
NOISE = ("קיצור דרך:", "קיצור דרך", "מסלול", "דרך")


def streets(name):
    """Best-effort split of a segment name into the endpoints it names."""
    cleaned = name
    for token in NOISE:
        if cleaned.startswith(token):
            cleaned = cleaned[len(token):]
    cleaned = cleaned.strip(" :")
    parts = [p.strip(" ()") for p in SPLIT.split(cleaned) if p.strip(" ()")]
    return [p for p in parts if len(p) > 1]


HEAD = PatternFill("solid", fgColor="1F4E79")
HEAD_FONT = Font(color="FFFFFF", bold=True)


def write_sheet(ws, rows, cols):
    ws.sheet_view.rightToLeft = True
    ws.append(cols)
    for cell in ws[1]:
        cell.fill = HEAD
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in rows:
        ws.append([row.get(c, "") for c in cols])
    widths = {"שם": 42, "הערה מהמפה": 55, "ניווט": 30, "רחובות בשם": 28, "סוג": 12}
    for i, col in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(col, 13)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def main():
    out = sys.argv[1] if len(sys.argv) > 1 else "derech_kitzur_map_data.xlsx"
    kml = requests.get(KML_URL, headers=H, timeout=60)
    kml.raise_for_status()
    rows = parse(kml.content)

    segs = sorted([r for r in rows if r["סוג"] == "מקטע שביל"], key=lambda r: -r["אורך (מ')"])
    pts = [r for r in rows if r["סוג"] == "נקודת ציון"]

    wb = Workbook()
    wb.remove(wb.active)

    write_sheet(
        wb.create_sheet("מקטעים"),
        segs,
        ["שם", "אורך (מ')", "מרחק אווירי (מ')", "פיתולים", "רחובות בשם",
         "הערה מהמפה", "תמונות", "נק' בקו", "קו רוחב", "קו אורך", "ניווט"],
    )
    write_sheet(
        wb.create_sheet("נקודות ציון"),
        pts,
        ["שם", "הערה מהמפה", "תמונות", "קו רוחב", "קו אורך", "ניווט"],
    )

    total = sum(r["אורך (מ')"] for r in segs)
    buckets = [("עד 50 מ'", 0, 50), ("51-100", 51, 100), ("101-200", 101, 200),
               ("201-300", 201, 300), ("מעל 300", 301, 10 ** 6)]
    summary = wb.create_sheet("סיכום")
    summary.sheet_view.rightToLeft = True
    summary.column_dimensions["A"].width = 34
    summary.column_dimensions["B"].width = 16
    for label, value in [
        ("מקור", MAP_URL),
        ("מספר מקטעי שביל", len(segs)),
        ("מספר נקודות ציון", len(pts)),
        ("אורך כולל (מ')", total),
        ("אורך כולל (ק\"מ)", round(total / 1000, 2)),
        ("מקטע ממוצע (מ')", round(total / len(segs)) if segs else 0),
        ("המקטע הקצר ביותר (מ')", min(r["אורך (מ')"] for r in segs) if segs else 0),
        ("המקטע הארוך ביותר (מ')", max(r["אורך (מ')"] for r in segs) if segs else 0),
        ("מקטעים עם הערה", sum(1 for r in segs if r["הערה מהמפה"])),
        ("מקטעים עם תמונה", sum(1 for r in segs if r["תמונות"])),
        ("מקטעים ללא הערה וללא תמונה",
         sum(1 for r in segs if not r["הערה מהמפה"] and not r["תמונות"])),
        ("", ""),
        ("התפלגות אורכים", "מקטעים"),
    ]:
        summary.append([label, value])
    for label, lo, hi in buckets:
        summary.append([label, sum(1 for r in segs if lo <= r["אורך (מ')"] <= hi)])
    for cell in ("A1", "A14"):
        summary[cell].font = Font(bold=True)

    wb.save(out)
    print(f"wrote {out}: {len(segs)} segments, {len(pts)} waypoints, {total} m")


if __name__ == "__main__":
    main()
